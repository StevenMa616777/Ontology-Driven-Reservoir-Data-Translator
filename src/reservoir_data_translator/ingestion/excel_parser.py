"""XLSX ingestion using openpyxl without semantic column interpretation."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import DocumentParser, IngestionError
from .models import RawBlock, RawDocument


class ExcelParser(DocumentParser):
    """Represent each non-empty worksheet as one rectangular table block."""

    source_type = "xlsx"
    suffixes = (".xlsx",)

    def parse(
        self,
        path: str | Path,
        *,
        source_id: str | None = None,
    ) -> RawDocument:
        source_path = self._source_path(path)
        try:
            workbook = load_workbook(
                source_path,
                read_only=True,
                data_only=False,
            )
        except Exception as exc:
            raise IngestionError(
                "MALFORMED_XLSX",
                f"Could not parse XLSX source {source_path.name}: {exc}",
                path=source_path,
            ) from exc

        blocks: list[RawBlock] = []
        try:
            for worksheet in workbook.worksheets:
                matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
                bounds = self._non_empty_bounds(matrix)
                if bounds is None:
                    continue
                first_row, last_row, first_column, last_column = bounds
                rectangle = [
                    row[first_column : last_column + 1]
                    for row in matrix[first_row : last_row + 1]
                ]
                columns = rectangle[0]
                rows = rectangle[1:]
                start = f"{get_column_letter(first_column + 1)}{first_row + 1}"
                end = f"{get_column_letter(last_column + 1)}{last_row + 1}"
                escaped_title = worksheet.title.replace("'", "''")
                blocks.append(
                    RawBlock(
                        block_id=f"block_{len(blocks) + 1:04d}",
                        block_type="table",
                        content={"columns": columns, "rows": rows},
                        source_location=f"'{escaped_title}'!{start}:{end}",
                    )
                )
        finally:
            workbook.close()

        return RawDocument(
            source_id=self._source_id(source_path, source_id),
            source_type=self.source_type,
            file_name=source_path.name,
            blocks=blocks,
        )

    @staticmethod
    def _non_empty_bounds(
        matrix: list[list[Any]],
    ) -> tuple[int, int, int, int] | None:
        populated = [
            (row_index, column_index)
            for row_index, row in enumerate(matrix)
            for column_index, value in enumerate(row)
            if value is not None
        ]
        if not populated:
            return None
        row_indices = [item[0] for item in populated]
        column_indices = [item[1] for item in populated]
        return (
            min(row_indices),
            max(row_indices),
            min(column_indices),
            max(column_indices),
        )


XlsxParser = ExcelParser
